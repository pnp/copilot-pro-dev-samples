"""
Builds a formatted Excel workbook from approved Dataverse extraction data.
Uses openpyxl. Returns workbook bytes.
"""
import io
import logging
import re
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# --- Style constants ---
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1a472a", end_color="1a472a", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color="1a472a")
SECTION_FILL = PatternFill(start_color="f1f0f7", end_color="f1f0f7", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)
TOTAL_BORDER = Border(top=Side(style="thin"), bottom=Side(style="double"))
NORMAL_FONT = Font(name="Calibri", size=10)
CORRECTED_FONT = Font(name="Calibri", size=10, color="6B4EE6")
ALT_ROW_FILL = PatternFill(start_color="fafaff", end_color="fafaff", fill_type="solid")
NUMBER_FORMAT = '#,##0.00'

STATEMENT_TYPE_MAP = {
    833060000: "Income Statement",
    833060001: "Balance Sheet",
    833060002: "Cash Flow",
}

ROW_TYPE_MAP = {
    833060000: "SectionHeader",
    833060001: "LineItem",
    833060002: "Subtotal",
    833060003: "Total",
}


def _get_display_value(item: dict) -> str | None:
    """Get the value to display: analyst correction takes precedence."""
    corrected = item.get("cree1_analystcorrectedvalue")
    if corrected:
        return corrected
    return item.get("cree1_valueraw")


def _try_numeric(val: str | None) -> float | str | None:
    """Try to parse a display value as a number for Excel formatting."""
    if val is None:
        return None
    cleaned = val.replace(",", "").replace("$", "").replace("(", "-").replace(")", "").strip()
    if not cleaned or cleaned == "\u2014":
        return None
    try:
        return float(cleaned)
    except ValueError:
        return val


def build_excel(job: dict, statements: list[dict], line_items: list[dict]) -> bytes:
    """Build an Excel workbook from Dataverse data and return as bytes."""
    wb = Workbook()

    company = job.get("cree1_companyname", "Financial Statements")

    # Group line items by statement record ID
    items_by_stmt: dict[str, list[dict]] = defaultdict(list)
    for item in line_items:
        stmt_id = item.get("_cree1_extractedstatement_value", "")
        if stmt_id:
            items_by_stmt[stmt_id].append(item)

    sheet_created = False

    for stmt in statements:
        stmt_id = stmt.get("cree1_extractedstatement1id", "")
        stmt_type_code = stmt.get("cree1_statementtype", 833060000)
        stmt_name = STATEMENT_TYPE_MAP.get(stmt_type_code, stmt.get("cree1_statementname", "Sheet"))
        stmt_items = items_by_stmt.get(stmt_id, [])

        if not stmt_items:
            continue

        if not sheet_created:
            ws = wb.active
            ws.title = stmt_name
            sheet_created = True
        else:
            ws = wb.create_sheet(title=stmt_name)

        _build_statement_sheet(ws, company, stmt_name, stmt_items, job)

    # Add summary sheet at the beginning
    if sheet_created:
        _build_summary_sheet(wb, job, statements, line_items)

    if not sheet_created:
        ws = wb.active
        ws.title = "No Data"
        ws["A1"] = "No extraction data available."

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_statement_sheet(ws, company: str, stmt_name: str, items: list[dict], job: dict | None = None):
    """Build a single statement worksheet."""
    # Determine unique periods (columns) and period types
    period_map: dict[int, str] = {}
    period_type_map: dict[int, str] = {}
    for item in items:
        col_idx = item.get("cree1_columnindex", 0)
        if col_idx not in period_map:
            period_map[col_idx] = item.get("cree1_period", "")
            period_type_map[col_idx] = item.get("cree1_periodtype", "")
    sorted_periods = sorted(period_map.items(), key=lambda x: x[0])

    # No delta columns in Excel (YoY/QoQ shown in grid only)
    delta_pairs: list[tuple[int, int, str]] = []

    num_period_cols = len(sorted_periods)
    num_delta_cols = 0

    # Title row
    total_cols = 1 + num_period_cols + num_delta_cols + 1  # label + periods + deltas + audit
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=f"Consolidated {stmt_name}")
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="1a472a")
    title_cell.alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 28

    # Currency/unit subtitle row
    currency = ""
    if job:
        cur = job.get("cree1_currencyname") or ""
        unit = job.get("cree1_currencyunitname") or ""
        if cur or unit:
            currency = f"{cur} in {unit.title()}" if cur and unit else cur or unit
    if currency:
        subtitle_cell = ws.cell(row=2, column=1, value=currency)
        subtitle_cell.font = Font(name="Calibri", size=10, italic=True, color="64748b")

    # Header row
    header_row = 4
    ws.cell(row=header_row, column=1, value="Line Item").font = HEADER_FONT
    ws.cell(row=header_row, column=1).fill = HEADER_FILL
    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal="left")

    # Build column layout: periods interleaved with delta columns
    col_layout: list[dict] = []  # list of {"type": "period"|"delta", ...}
    delta_idx = 0
    for p_idx, (col_idx, period_label) in enumerate(sorted_periods):
        col_layout.append({"type": "period", "col_idx": col_idx, "label": period_label})
        # Check if this period starts a delta pair
        if delta_idx < len(delta_pairs) and delta_pairs[delta_idx][0] == col_idx:
            col_layout.append({"type": "delta", "current": col_idx, "prior": delta_pairs[delta_idx][1], "label": delta_pairs[delta_idx][2]})
            delta_idx += 1

    for i, col_def in enumerate(col_layout):
        cell = ws.cell(row=header_row, column=2 + i, value=col_def["label"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="right")

    # Audit column header
    audit_col = 2 + len(col_layout)
    audit_cell = ws.cell(row=header_row, column=audit_col, value="Corrections")
    audit_cell.font = HEADER_FONT
    audit_cell.fill = HEADER_FILL
    audit_cell.alignment = Alignment(horizontal="left")

    # Group items by row index
    rows_by_idx: dict[int, dict[int, dict]] = defaultdict(dict)
    row_meta: dict[int, dict] = {}
    for item in items:
        ridx = item.get("cree1_rowindex", 0)
        cidx = item.get("cree1_columnindex", 0)
        rows_by_idx[ridx][cidx] = item
        if ridx not in row_meta:
            row_meta[ridx] = item

    sorted_row_indices = sorted(rows_by_idx.keys())

    current_row = header_row + 1
    for data_row_num, ridx in enumerate(sorted_row_indices):
        meta = row_meta[ridx]
        row_type = ROW_TYPE_MAP.get(meta.get("cree1_rowtype", 833060001), "LineItem")
        label = meta.get("cree1_lineitemname", "") or meta.get("cree1_labelraw", "")
        label = re.sub(r'\s*[（(](?:Loss|Losses?)\s+(?:shown|indicated)\s+as\s+["\'-]+[)）]', "", label, flags=re.IGNORECASE)
        indent = meta.get("cree1_indentlevel", 0)

        # Label cell
        label_cell = ws.cell(row=current_row, column=1, value=label)
        label_cell.alignment = Alignment(indent=indent)

        total_data_cols = len(col_layout)
        if row_type == "SectionHeader":
            label_cell.font = SECTION_FONT
            for c in range(1, 2 + total_data_cols + 1):
                ws.cell(row=current_row, column=c).fill = SECTION_FILL
        elif row_type in ("Total", "Subtotal"):
            label_cell.font = TOTAL_FONT
            for c in range(1, 2 + total_data_cols):
                ws.cell(row=current_row, column=c).border = TOTAL_BORDER
        else:
            label_cell.font = NORMAL_FONT
            if data_row_num % 2 == 1:
                for c in range(1, 2 + total_data_cols + 1):
                    ws.cell(row=current_row, column=c).fill = ALT_ROW_FILL

        # Value cells + delta cells + audit trail
        corrections = []
        for i, col_def in enumerate(col_layout):
            value_cell = ws.cell(row=current_row, column=2 + i)

            if row_type == "SectionHeader":
                continue

            if col_def["type"] == "period":
                col_idx = col_def["col_idx"]
                item = rows_by_idx[ridx].get(col_idx)
                if item:
                    display = _get_display_value(item)
                    numeric = _try_numeric(display)

                    if isinstance(numeric, float):
                        value_cell.value = numeric
                        value_cell.number_format = NUMBER_FORMAT
                    elif numeric is not None:
                        value_cell.value = numeric
                    else:
                        value_cell.value = "\u2014"

                    value_cell.alignment = Alignment(horizontal="right")

                    corrected = item.get("cree1_analystcorrectedvalue")
                    original = item.get("cree1_valueraw")
                    is_corrected = corrected and corrected != original
                    if is_corrected:
                        value_cell.font = CORRECTED_FONT
                        period_label = period_map.get(col_idx, "")
                        corrections.append(f"{period_label}: {original} -> {corrected}")
                    elif row_type in ("Total", "Subtotal"):
                        value_cell.font = TOTAL_FONT
                    else:
                        value_cell.font = NORMAL_FONT

            elif col_def["type"] == "delta":
                curr_item = rows_by_idx[ridx].get(col_def["current"])
                prior_item = rows_by_idx[ridx].get(col_def["prior"])
                curr_val = curr_item.get("cree1_valuenormalized") if curr_item else None
                prior_val = prior_item.get("cree1_valuenormalized") if prior_item else None

                if curr_val is not None and prior_val is not None:
                    try:
                        delta = float(curr_val) - float(prior_val)
                        value_cell.value = delta
                        value_cell.number_format = NUMBER_FORMAT
                        value_cell.alignment = Alignment(horizontal="right")
                        if delta < 0:
                            value_cell.font = Font(name="Calibri", size=10, color="dc2626")
                        else:
                            value_cell.font = Font(name="Calibri", size=10, color="16a34a")
                    except (TypeError, ValueError):
                        pass

        # Write audit column
        if corrections:
            audit_cell = ws.cell(row=current_row, column=audit_col, value="; ".join(corrections))
            audit_cell.font = Font(name="Calibri", size=9, color="6B4EE6", italic=True)

        current_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 40
    for i in range(len(col_layout)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18
    ws.column_dimensions[get_column_letter(audit_col)].width = 35


def _build_summary_sheet(wb: Workbook, job: dict, statements: list[dict], line_items: list[dict]):
    """Add a Summary sheet at the beginning of the workbook."""
    ws = wb.create_sheet("Summary", 0)

    company = job.get("cree1_companyname", "")
    ws.cell(row=1, column=1, value=f"{company} — Extraction Summary")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="1a472a")

    # --- Extraction metadata ---
    meta_rows = [
        ("Job ID", job.get("cree1_jobid", "")),
        ("File", job.get("cree1_filename", "")),
        ("Statements Found", job.get("cree1_statementsfound", "")),
        ("Total Line Items", job.get("cree1_totallineitems", "")),
        ("Avg Confidence", f"{(job.get('cree1_avgconfidence') or 0) * 100:.1f}%"),
        ("Status", "Approved"),
    ]

    for i, (label, value) in enumerate(meta_rows):
        ws.cell(row=3 + i, column=1, value=label).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=3 + i, column=2, value=str(value)).font = Font(name="Calibri", size=10)

    # --- Extraction status table (which statements, pages, status) ---
    table_start = 3 + len(meta_rows) + 1
    ws.cell(row=table_start, column=1, value="Extraction Status").font = Font(name="Calibri", size=12, bold=True, color="1a472a")

    header_row = table_start + 1
    for col_idx, header in enumerate(["Financial Statement", "Source Pages", "Extraction Status", "Coverage", "Notes"], 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    # Build statement status rows — keyed on integer type code so lookups below work correctly
    stmt_by_type: dict[int, dict] = {}
    for s in statements:
        code = s.get("cree1_statementtype")
        if code is not None:
            stmt_by_type[code] = s

    # Count line items per statement
    items_per_stmt: dict[str, int] = {}
    for item in line_items:
        sid = item.get("_cree1_extractedstatement_value", "")
        items_per_stmt[sid] = items_per_stmt.get(sid, 0) + 1

    stmt_display = [
        (833060001, "Balance Sheet"),
        (833060000, "Income Statement"),
        (833060002, "Cash Flow"),
    ]

    current_row = header_row + 1
    for type_code, display_name in stmt_display:
        stmt = stmt_by_type.get(type_code)
        ws.cell(row=current_row, column=1, value=display_name).font = NORMAL_FONT

        if stmt:
            # Pages
            ps = stmt.get("cree1_pagerangestart")
            pe = stmt.get("cree1_pagerangeend")
            if ps and pe:
                page_text = f"Pages {ps}-{pe}" if ps != pe else f"Page {ps}"
            elif ps:
                page_text = f"Page {ps}"
            else:
                page_text = "—"
            ws.cell(row=current_row, column=2, value=page_text).font = NORMAL_FONT

            # Status
            ws.cell(row=current_row, column=3, value="Extracted").font = Font(name="Calibri", size=10, color="16a34a")

            # Coverage
            stmt_id = stmt.get("cree1_extractedstatement1id", "")
            item_count = items_per_stmt.get(stmt_id, 0)
            ws.cell(row=current_row, column=4, value=f"{item_count} line items").font = NORMAL_FONT

            # Notes
            title = stmt.get("cree1_statementtitle", "")
            if title:
                ws.cell(row=current_row, column=5, value=title).font = Font(name="Calibri", size=10, italic=True, color="64748b")
        else:
            ws.cell(row=current_row, column=2, value="—").font = NORMAL_FONT
            ws.cell(row=current_row, column=3, value="Not found").font = Font(name="Calibri", size=10, color="dc2626")
            ws.cell(row=current_row, column=4, value="—").font = NORMAL_FONT

        current_row += 1

    # --- Analyst corrections ---
    current_row += 1
    corrected = [item for item in line_items if item.get("cree1_analystcorrectedvalue")]
    ws.cell(row=current_row, column=1, value="Analyst Corrections").font = Font(name="Calibri", size=12, bold=True, color="6B4EE6")
    current_row += 1

    if corrected:
        for col_idx, header in enumerate(["Line Item", "Period", "Original", "Corrected"], 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        current_row += 1

        for item in corrected:
            ws.cell(row=current_row, column=1, value=item.get("cree1_lineitemname", "")).font = NORMAL_FONT
            ws.cell(row=current_row, column=2, value=item.get("cree1_period", "")).font = NORMAL_FONT
            ws.cell(row=current_row, column=3, value=item.get("cree1_valueraw", "")).font = NORMAL_FONT
            ws.cell(row=current_row, column=4, value=item.get("cree1_analystcorrectedvalue", "")).font = CORRECTED_FONT
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="No corrections made.").font = Font(name="Calibri", size=10, italic=True, color="64748b")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 50
