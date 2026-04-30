"""
Professional Excel formatter for financial statement extraction output.

Generates beautifully formatted Excel workbooks matching institutional
financial reporting standards:
  - Dark navy header bars with white text
  - Teal subtitle rows with period/currency info
  - Bold italic section headers with bilingual labels
  - Red negative values
  - Color-coded sheet tabs
  - Freeze panes on header rows
  - Conditional number formatting
"""
import re
import openpyxl.workbook.properties
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter

from .stages.enrich import translate_column_header


# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------

NAVY = "1B2A4A"
TEAL = "2E6B77"
DARK_HEADER = "333333"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
RED = "CC0000"
LIGHT_BLUE = "DCE6F1"

# Sheet tab colors
TAB_COLORS = {
    "balance_sheet": "339966",       # Green
    "income_statement": "CC3333",    # Red
    "cash_flow": "E67300",           # Orange
}

# ---------------------------------------------------------------------------
# Reusable styles
# ---------------------------------------------------------------------------

FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color=WHITE)
FONT_COL_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_SECTION = Font(name="Calibri", size=11, bold=True, italic=True)
FONT_NORMAL = Font(name="Calibri", size=11)
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_RED = Font(name="Calibri", size=11, color=RED)
FONT_RED_BOLD = Font(name="Calibri", size=11, bold=True, color=RED)
FONT_SUMMARY_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_NOTES = Font(name="Calibri", size=10, italic=True)

FILL_NAVY = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
FILL_TEAL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
FILL_DARK = PatternFill(start_color=DARK_HEADER, end_color=DARK_HEADER, fill_type="solid")
FILL_LIGHT_BLUE = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

BORDER_BOTTOM = Border(bottom=Side(style="thin"))
BORDER_TOP_BOTTOM = Border(top=Side(style="thin"), bottom=Side(style="double"))

NUMBER_FORMAT = '#,##0.00'


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _build_summary_sheet(ws, result: dict, title: str, fx_config: dict | None = None):
    """Build a professional summary sheet."""
    summary = result.get("summary", [])

    # Title bar (merged across 4 columns)
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = title
    cell.font = FONT_TITLE
    cell.fill = FILL_NAVY
    cell.alignment = ALIGN_CENTER
    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = FILL_NAVY

    # Subtitle
    ws.merge_cells("A2:D2")
    # Build subtitle from first statement's metadata
    subtitle_parts = []
    for stype in ["balance_sheet", "income_statement", "cash_flow"]:
        doc = result.get(stype)
        if doc and isinstance(doc, dict):
            sm = doc.get("statement_metadata", {})
            if sm.get("currency"):
                subtitle_parts.append(f"Currency: {sm['currency']}")
            break
    cell = ws["A2"]
    cell.value = " | ".join(subtitle_parts) if subtitle_parts else ""
    cell.font = FONT_SUBTITLE
    cell.fill = FILL_TEAL
    cell.alignment = ALIGN_CENTER
    for col in range(1, 5):
        ws.cell(row=2, column=col).fill = FILL_TEAL

    # Column headers
    headers = ["Statement", "Description", "Pages in Report", "Sheet Reference"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = FONT_SUMMARY_HEADER
        cell.fill = FILL_DARK
        cell.alignment = ALIGN_CENTER

    # Statement rows
    STYPE_NAMES = {
        "balance_sheet": ("Consolidated Balance Sheet", "Assets, liabilities and equity"),
        "income_statement": ("Consolidated Income Statement", "Revenue and profit"),
        "cash_flow": ("Consolidated Cash Flow", "Cash flows"),
    }

    row = 4
    for s in summary:
        stype = s.get("statement_type", "")
        names = STYPE_NAMES.get(stype, (stype, ""))
        pr = s.get("page_range", {})
        pages = f"{pr.get('start', '?')}-{pr.get('end', '?')}" if pr.get("start") else "-"
        quality = s.get("quality_score")

        ws.cell(row=row, column=1, value=names[0]).font = FONT_NORMAL
        ws.cell(row=row, column=2, value=names[1]).font = FONT_NORMAL
        c_pages = ws.cell(row=row, column=3, value=pages)
        c_pages.font = FONT_NORMAL
        c_pages.alignment = ALIGN_CENTER

        sheet_ref = names[0].split()[-2] + " " + names[0].split()[-1] if len(names[0].split()) > 2 else names[0]
        ws.cell(row=row, column=4, value=sheet_ref).font = FONT_NORMAL
        ws.cell(row=row, column=4).alignment = ALIGN_CENTER
        row += 1

    # Notes section
    row += 1
    ws.cell(row=row, column=1, value="Notes:").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=row, column=1).fill = FILL_LIGHT_BLUE
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = FILL_LIGHT_BLUE
    row += 1

    # Get currency info from first available statement
    currency = ""
    currency_symbol = ""
    unit = ""
    for stype in ["balance_sheet", "income_statement", "cash_flow"]:
        doc = result.get(stype)
        if doc and isinstance(doc, dict):
            sm = doc.get("statement_metadata", {})
            currency = sm.get("currency", "")
            currency_symbol = sm.get("currency_symbol", "")
            unit = sm.get("unit", "")
            break

    # Currency and unit
    currency_display = f"{currency} ({currency_symbol})" if currency_symbol else currency
    if currency:
        ws.cell(row=row, column=1, value=f"Reporting Currency: {currency_display}").font = FONT_NOTES
        row += 1
    if unit:
        ws.cell(row=row, column=1, value=f"Unit: {unit.capitalize()}").font = FONT_NOTES
        row += 1

    # FX methodology statement
    ws.cell(row=row, column=1,
            value=f"All figures are presented in {currency_display or 'the original reporting currency'} "
                  f"as reported in the source document. No foreign exchange conversion has been applied."
            ).font = FONT_NOTES
    row += 1

    # FX methodology statement
    if fx_config:
        sc = fx_config["source_currency"]
        tc = fx_config["target_currency"]
        spot = fx_config["spot_rate"]
        avg = fx_config["avg_rate"]
        rd = fx_config.get("rate_date", "")
        src = fx_config.get("rate_source", "")
        ws.cell(row=row, column=1,
                value=f"FX Conversion: Balance Sheet items converted at closing spot rate "
                      f"{sc}/{tc} {spot:.4f} as at {rd}. Income Statement and Cash Flow items "
                      f"converted at period average rate {sc}/{tc} {avg:.4f}. "
                      f"Source: {src}. Rates are editable in each sheet — modify the rate cell to recalculate."
                ).font = FONT_NOTES
        row += 1
    else:
        ws.cell(row=row, column=1,
                value=f"All figures are presented in {currency_display or 'the original reporting currency'} "
                      f"as reported in the source document. No foreign exchange conversion has been applied."
                ).font = FONT_NOTES
        row += 1

    # Analyst corrections note
    corrections_count = 0
    for stype in ["balance_sheet", "income_statement", "cash_flow"]:
        doc = result.get(stype)
        if doc and isinstance(doc, dict):
            for r in doc.get("rows", []):
                for v in r.get("values", []):
                    if v.get("corrected"):
                        corrections_count += 1
    if corrections_count > 0:
        ws.cell(row=row, column=1,
                value=f"Analyst Corrections: {corrections_count} value(s) were manually adjusted during HITL review."
                ).font = Font(name="Calibri", size=10, italic=True, bold=True, color="CC3333")
        row += 1

    ws.cell(row=row, column=1, value="All figures as reported; no rounding applied.").font = FONT_NOTES
    row += 1
    ws.cell(row=row, column=1, value="Prepared by Microsoft Copilot from source PDF.").font = FONT_NOTES

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20


# ---------------------------------------------------------------------------
# Statement sheet
# ---------------------------------------------------------------------------

def _build_statement_sheet(ws, doc: dict, stype: str, fx_config: dict | None = None):
    """Build a professionally formatted statement sheet."""
    sm = doc.get("statement_metadata", {})
    cols = doc.get("columns", [])
    rows = doc.get("rows", [])

    # Extract currency info early (needed for FX column headers)
    currency = sm.get("currency", "")
    currency_symbol = sm.get("currency_symbol", "")
    unit = sm.get("unit", "")

    # Determine value columns (skip label column)
    value_col_headers = []
    for col in cols:
        label = col.get("label", "")
        translated = translate_column_header(label)
        if translated.lower() in ("item", "") or label == "項目":
            continue
        value_col_headers.append(translated)

    # If FX conversion, double the value columns (original + converted)
    if fx_config:
        fx_col_headers = []
        for h in value_col_headers:
            fx_col_headers.append(f"{h} ({currency})")
            fx_col_headers.append(f"{h} ({fx_config['target_currency']})")
        display_col_headers = fx_col_headers
    else:
        display_col_headers = value_col_headers

    total_cols = 2 + len(display_col_headers)  # Chinese + English + value cols
    last_col_letter = get_column_letter(total_cols)

    # Row 1: Title bar
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_raw = sm.get("statement_title_raw", "")
    title_en = sm.get("statement_title", "")
    title_text = f"{title_en} ({title_raw})" if title_raw and title_raw != title_en else title_en
    cell = ws["A1"]
    cell.value = title_text
    cell.font = FONT_TITLE
    cell.fill = FILL_NAVY
    cell.alignment = ALIGN_CENTER
    for ci in range(1, total_cols + 1):
        ws.cell(row=1, column=ci).fill = FILL_NAVY

    # Row 2: Subtitle
    ws.merge_cells(f"A2:{last_col_letter}2")
    currency_display = f"{currency} ({currency_symbol})" if currency_symbol else currency

    if fx_config:
        tc = fx_config["target_currency"]
        rate = fx_config["applied_rate"]
        rate_type = fx_config["rate_type"]
        subtitle = f"Original: {currency_display} | Converted to: {tc} | Rate: {rate:.4f} ({rate_type}) | Unit: {unit.capitalize()}"
    else:
        subtitle = f"Reporting Currency: {currency_display} | Unit: {unit.capitalize()}" if currency else ""
    cell = ws["A2"]
    cell.value = subtitle
    cell.font = FONT_SUBTITLE
    cell.fill = FILL_TEAL
    cell.alignment = ALIGN_CENTER
    for ci in range(1, total_cols + 1):
        ws.cell(row=2, column=ci).fill = FILL_TEAL

    # Row 3: Column headers
    col_headers = ["Item (Chinese)", "Item (English)"] + display_col_headers
    for ci, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_DARK
        cell.alignment = ALIGN_CENTER

    # FX rate cell (editable by analyst) — placed to the right of data columns
    rate_cell_col = total_cols + 2  # 2 columns gap after data
    rate_cell_row = 2
    if fx_config:
        rate_label_cell = ws.cell(row=1, column=rate_cell_col, value="FX Rate →")
        rate_label_cell.font = Font(name="Calibri", size=10, bold=True, color="CC3333")
        rate_cell = ws.cell(row=rate_cell_row, column=rate_cell_col, value=fx_config["applied_rate"])
        rate_cell.font = Font(name="Calibri", size=12, bold=True, color="CC3333")
        rate_cell.number_format = "0.000000"
        rate_cell_ref = f"${get_column_letter(rate_cell_col)}${rate_cell_row}"

    # Freeze panes below header
    ws.freeze_panes = "A4"

    # Data rows
    for ri, row in enumerate(rows, start=4):
        label_raw = row.get("label_raw", "")
        label_norm = row.get("label_normalized") or label_raw
        row_type = row.get("row_type", "line_item")
        indent = row.get("indent_level", 0)

        # Clean label noise
        label_norm = re.sub(
            r'\s*[(\uff08](?:Loss|Losses?)\s+(?:shown|indicated)\s+as.*?[)\uff09]',
            '', label_norm, flags=re.IGNORECASE,
        )

        # Section header: bold italic, bilingual, merged look
        if row_type == "section_header":
            display = f"{label_norm.upper()} ({label_raw})" if label_raw and label_raw != label_norm else label_norm.upper()
            cell_a = ws.cell(row=ri, column=1, value=display)
            cell_a.font = FONT_SECTION
            ws.cell(row=ri, column=2).value = ""
            continue

        # Chinese label
        cell_a = ws.cell(row=ri, column=1, value=label_raw)
        cell_a.font = FONT_NORMAL
        if indent > 0:
            cell_a.alignment = Alignment(indent=indent * 2)

        # English label
        is_total = row_type in ("subtotal", "total", "grand_total")
        cell_b = ws.cell(row=ri, column=2, value=label_norm.upper() if is_total else label_norm)
        cell_b.font = FONT_BOLD if is_total else FONT_NORMAL
        if indent > 0 and not is_total:
            cell_b.alignment = Alignment(indent=indent * 2)

        # Add border for totals
        if is_total:
            for ci in range(1, total_cols + 1):
                if row_type == "grand_total":
                    ws.cell(row=ri, column=ci).border = BORDER_TOP_BOTTOM
                else:
                    ws.cell(row=ri, column=ci).border = BORDER_BOTTOM

        # Values
        values = row.get("values", [])
        if fx_config:
            # Dual columns: original (C, E, ...) + converted formula (D, F, ...)
            col_offset = 3
            for vi in range(min(len(values), len(value_col_headers))):
                val = values[vi]
                normalized = val.get("normalized")
                raw_val = val.get("raw")

                # Original value column
                orig_col = col_offset
                cell = ws.cell(row=ri, column=orig_col)
                cell.alignment = ALIGN_RIGHT
                if normalized is not None:
                    cell.value = normalized
                    cell.number_format = NUMBER_FORMAT
                    if normalized < 0:
                        cell.font = FONT_RED_BOLD if is_total else FONT_RED
                    else:
                        cell.font = FONT_BOLD if is_total else FONT_NORMAL
                elif raw_val:
                    cell.value = raw_val
                    cell.font = FONT_BOLD if is_total else FONT_NORMAL

                # Converted value column (Excel formula referencing rate cell)
                fx_col = col_offset + 1
                fx_cell = ws.cell(row=ri, column=fx_col)
                fx_cell.alignment = ALIGN_RIGHT
                fx_cell.number_format = NUMBER_FORMAT
                if normalized is not None:
                    orig_ref = f"{get_column_letter(orig_col)}{ri}"
                    fx_cell.value = f"={orig_ref}*{rate_cell_ref}"
                    fx_cell.font = FONT_BOLD if is_total else FONT_NORMAL
                    # Apply light blue background to converted columns
                    fx_cell.fill = FILL_LIGHT_BLUE if not is_total else PatternFill()

                col_offset += 2  # Skip 2 (original + converted)
        else:
            # No FX: single value columns
            for vi in range(min(len(values), len(value_col_headers))):
                val = values[vi]
                normalized = val.get("normalized")
                raw_val = val.get("raw")
                cell = ws.cell(row=ri, column=3 + vi)
                cell.alignment = ALIGN_RIGHT

                if normalized is not None:
                    cell.value = normalized
                    cell.number_format = NUMBER_FORMAT
                    if normalized < 0:
                        cell.font = FONT_RED_BOLD if is_total else FONT_RED
                    else:
                        cell.font = FONT_BOLD if is_total else FONT_NORMAL
                elif raw_val:
                    cell.value = raw_val
                    cell.font = FONT_BOLD if is_total else FONT_NORMAL

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    for ci in range(len(display_col_headers)):
        ws.column_dimensions[get_column_letter(3 + ci)].width = 22

    # Tab color
    if stype in TAB_COLORS:
        ws.sheet_properties.tabColor = TAB_COLORS[stype]


# ---------------------------------------------------------------------------
# Margins & Ratios sheet
# ---------------------------------------------------------------------------

# Canonical key aliases used to locate values across BS / IS / CF rows
_KEY_ALIASES = {
    "revenue": ["total_operating_revenue", "revenue", "net_revenue"],
    "cogs": ["cost_of_goods_sold", "operating_costs", "cost_of_revenue"],
    "operating_profit": ["operating_profit", "operating_income"],
    "net_profit": ["net_income", "net_profit", "profit_for_the_period"],
    "total_assets": ["total_assets"],
    "total_equity": ["total_equity", "total_owners_equity"],
    "total_liabilities": ["total_liabilities"],
    "current_assets": ["total_current_assets"],
    "current_liabilities": ["total_current_liabilities"],
    "inventories": ["inventories", "inventory"],
}

# Ratio definitions organised by section
_RATIO_SECTIONS = [
    ("PROFITABILITY", [
        ("Gross Margin (%)", "pct", lambda v: _safe_div(v["revenue"] - v["cogs"], v["revenue"]) * 100,
         ["revenue", "cogs"]),
        ("EBIT Margin (%)", "pct", lambda v: _safe_div(v["operating_profit"], v["revenue"]) * 100,
         ["operating_profit", "revenue"]),
        ("Net Profit Margin (%)", "pct", lambda v: _safe_div(v["net_profit"], v["revenue"]) * 100,
         ["net_profit", "revenue"]),
        ("Return on Equity (ROE) (%)", "pct", lambda v: _safe_div(v["net_profit"], v["total_equity"]) * 100,
         ["net_profit", "total_equity"]),
        ("Return on Assets (ROA) (%)", "pct", lambda v: _safe_div(v["net_profit"], v["total_assets"]) * 100,
         ["net_profit", "total_assets"]),
    ]),
    ("LIQUIDITY", [
        ("Current Ratio", "ratio", lambda v: _safe_div(v["current_assets"], v["current_liabilities"]),
         ["current_assets", "current_liabilities"]),
        ("Quick Ratio", "ratio",
         lambda v: _safe_div(v["current_assets"] - v.get("inventories", 0), v["current_liabilities"]),
         ["current_assets", "current_liabilities"]),
    ]),
    ("LEVERAGE", [
        ("Debt to Equity", "ratio", lambda v: _safe_div(v["total_liabilities"], v["total_equity"]),
         ["total_liabilities", "total_equity"]),
        ("Total Debt Ratio", "ratio", lambda v: _safe_div(v["total_liabilities"], v["total_assets"]),
         ["total_liabilities", "total_assets"]),
    ]),
]


def _safe_div(numerator, denominator):
    """Return numerator / denominator, or None when division is impossible."""
    if numerator is None or denominator is None or denominator == 0:
        return None
    return numerator / denominator


def _collect_values(result: dict):
    """Scan BS, IS, CF rows and return {period_index: {key: value}} maps.

    Each period_index corresponds to a column in the source data (0, 1, …).
    Returns (period_values_dict, period_labels_list).
    """
    # Gather all rows from all statement types
    all_rows = []
    for stype in ["balance_sheet", "income_statement", "cash_flow"]:
        doc = result.get(stype)
        if doc and isinstance(doc, dict):
            all_rows.extend(doc.get("rows", []))

    # Determine period labels from first available statement's columns
    period_labels: list[str] = []
    for stype in ["income_statement", "balance_sheet", "cash_flow"]:
        doc = result.get(stype)
        if doc and isinstance(doc, dict):
            cols = doc.get("columns", [])
            for col in cols:
                label = col.get("label", "")
                translated = translate_column_header(label)
                if translated.lower() in ("item", "") or label == "項目":
                    continue
                period_labels.append(translated)
            if period_labels:
                break

    # Build per-period value maps  {period_idx: {metric_key: float}}
    period_values: dict[int, dict[str, float]] = {}
    for row in all_rows:
        canonical = row.get("canonical_key", "")
        if not canonical:
            continue
        for alias_key, aliases in _KEY_ALIASES.items():
            if canonical in aliases:
                values = row.get("values", [])
                for vi, val in enumerate(values):
                    normalized = val.get("normalized")
                    if normalized is not None:
                        period_values.setdefault(vi, {})[alias_key] = normalized
                break  # matched alias group

    return period_values, period_labels


def _format_ratio(value, fmt: str) -> str:
    """Format a computed ratio value for display."""
    if value is None:
        return "N/A"
    if fmt == "pct":
        return f"{value:.1f}%"
    return f"{value:.2f}x"


def _build_margins_sheet(ws, result: dict):
    """Build a Margins & Ratios computed sheet from extracted data."""
    period_values, period_labels = _collect_values(result)

    # If we have no period labels, create generic ones
    if not period_labels:
        if period_values:
            period_labels = [f"Period {i+1}" for i in sorted(period_values.keys())]
        else:
            period_labels = ["Period 1"]

    num_periods = len(period_labels)
    total_cols = 1 + num_periods  # Metric + period columns
    last_col_letter = get_column_letter(total_cols)

    # Row 1: Title bar
    ws.merge_cells(f"A1:{last_col_letter}1")
    cell = ws["A1"]
    cell.value = "Margins & Ratios"
    cell.font = FONT_TITLE
    cell.fill = FILL_NAVY
    cell.alignment = ALIGN_CENTER
    for ci in range(1, total_cols + 1):
        ws.cell(row=1, column=ci).fill = FILL_NAVY

    # Row 2: Subtitle
    ws.merge_cells(f"A2:{last_col_letter}2")
    cell = ws["A2"]
    cell.value = "Computed from extracted financial statements"
    cell.font = FONT_SUBTITLE
    cell.fill = FILL_TEAL
    cell.alignment = ALIGN_CENTER
    for ci in range(1, total_cols + 1):
        ws.cell(row=2, column=ci).fill = FILL_TEAL

    # Row 3: Column headers
    col_headers = ["Metric"] + period_labels
    for ci, h in enumerate(col_headers, start=1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = FONT_COL_HEADER
        cell.fill = FILL_DARK
        cell.alignment = ALIGN_CENTER

    ws.freeze_panes = "A4"

    # Data rows
    current_row = 4
    for section_name, ratios in _RATIO_SECTIONS:
        # Section header
        cell = ws.cell(row=current_row, column=1, value=section_name)
        cell.font = FONT_SECTION
        cell.fill = FILL_LIGHT_GRAY
        for ci in range(2, total_cols + 1):
            ws.cell(row=current_row, column=ci).fill = FILL_LIGHT_GRAY
        current_row += 1

        for label, fmt, compute_fn, required_keys in ratios:
            ws.cell(row=current_row, column=1, value=label).font = FONT_NORMAL

            for pi in range(num_periods):
                pv = period_values.get(pi, {})
                # Check all required keys are present
                has_all = all(k in pv for k in required_keys)
                if has_all:
                    try:
                        value = compute_fn(pv)
                    except Exception:
                        value = None
                else:
                    value = None

                display = _format_ratio(value, fmt)
                cell = ws.cell(row=current_row, column=2 + pi, value=display)
                cell.alignment = ALIGN_RIGHT
                if display == "N/A":
                    cell.font = Font(name="Calibri", size=11, italic=True, color="999999")
                elif fmt == "pct" and value is not None and value < 0:
                    cell.font = FONT_RED
                else:
                    cell.font = FONT_NORMAL

            current_row += 1

        # Blank row between sections
        current_row += 1

    # Column widths
    ws.column_dimensions["A"].width = 35
    for ci in range(num_periods):
        ws.column_dimensions[get_column_letter(2 + ci)].width = 22

    # Purple tab color
    ws.sheet_properties.tabColor = "6B4EE6"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_professional_excel(
    result: dict,
    output_path: str,
    title: str = "Financial Statement Extraction",
    fx_target_currency: str | None = None,
    fx_spot_rate: float | None = None,
    fx_avg_rate: float | None = None,
    fx_rate_date: str | None = None,
    fx_rate_source: str = "exchangerate.host",
):
    """Generate a professionally formatted Excel workbook from extraction result.

    Args:
        result: Pipeline output dict with summary + statement docs.
        output_path: Path to save the .xlsx file.
        title: Title for the summary sheet header.
        fx_target_currency: Target currency for FX conversion (e.g., "AUD"). None = no conversion.
        fx_spot_rate: Closing spot rate for BS items.
        fx_avg_rate: Period average rate for IS/CF items.
        fx_rate_date: Date of the FX rate.
        fx_rate_source: Source of the rate (e.g., "exchangerate.host", "manual").
    """
    fx_config = None
    if fx_target_currency and fx_spot_rate:
        # Get source currency from first statement
        source_currency = ""
        for stype in ["balance_sheet", "income_statement", "cash_flow"]:
            doc = result.get(stype)
            if doc and isinstance(doc, dict):
                source_currency = doc.get("statement_metadata", {}).get("currency", "")
                if source_currency:
                    break
        fx_config = {
            "source_currency": source_currency,
            "target_currency": fx_target_currency,
            "spot_rate": fx_spot_rate,
            "avg_rate": fx_avg_rate or fx_spot_rate,
            "rate_date": fx_rate_date or "",
            "rate_source": fx_rate_source,
        }

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _build_summary_sheet(ws_summary, result, title, fx_config=fx_config)

    # Statement sheets
    SHEET_NAMES = {
        "balance_sheet": "Balance Sheet",
        "income_statement": "Income Statement",
        "cash_flow": "Cash Flow",
    }

    # Rate per statement type: BS uses spot, IS/CF use average
    RATE_BY_STYPE = {
        "balance_sheet": "spot_rate",
        "income_statement": "avg_rate",
        "cash_flow": "avg_rate",
    }

    for stype, sheet_name in SHEET_NAMES.items():
        doc = result.get(stype)
        if not doc or not isinstance(doc, dict):
            continue
        ws = wb.create_sheet(title=sheet_name)
        rate_key = RATE_BY_STYPE[stype]
        stype_fx = fx_config.copy() if fx_config else None
        if stype_fx:
            stype_fx["applied_rate"] = stype_fx[rate_key]
            stype_fx["rate_type"] = "closing spot" if rate_key == "spot_rate" else "period average"
        _build_statement_sheet(ws, doc, stype, fx_config=stype_fx)

    # Margins & Ratios computed sheet
    ws_margins = wb.create_sheet(title="Margins & Ratios")
    _build_margins_sheet(ws_margins, result)

    # Force Excel to recalculate all formulas on open
    wb.calculation = openpyxl.workbook.properties.CalcProperties(fullCalcOnLoad=True)

    wb.save(output_path)
    return output_path
